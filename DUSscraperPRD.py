import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Webscraper
page = 1  # Start from page 13 for speed
URL = "http://dusunnah.com/articles/page/"
all_siteBLOGS = []
while True:
    site = requests.get(URL + str(page))
    # siteurl = site.url
    # print(siteurl)

    page = page+1
    siteSRC = site.content
    # Creates Soup Object
    soup = BeautifulSoup(siteSRC, 'lxml')
    # Find class that each article uses
    siteBLOGS = soup.find_all(class_="blog-info")
    if len(siteBLOGS) == 0:
        break

    all_siteBLOGS.extend(siteBLOGS)


def output_to_xlsx():
    # Create Excel spreadsheet
    file = Workbook()
    ws = file.active
    ws.title = "All Dus Articles"
    # Write to rows
    for index, article in enumerate(all_siteBLOGS):
        ws.cell(row=index + 1, column=1).value = article.a.text
        ws.cell(row=index + 1, column=2).value = article.a['href']
        ws.row_dimensions[index + 1].height = 28
        ws.column_dimensions['A'].width = 120
        ws.column_dimensions['B'].width = 180
    # Save spreadsheet
    file.save('DUS ARTICLES.xlsx')


output_to_xlsx()
