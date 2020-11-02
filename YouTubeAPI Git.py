# Import key modules
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font

# Create build with API key
api_key = "" # Enter Google API Key
youtube = build('youtube', 'v3', developerKey=api_key)
# Channel & uploads playlist all_video_ids
all_uploads = "" # Enter All Uploads Playlist ID
Channel_id = "" # Enter Channel ID
pagetoken = None
# Video Title and Id lists
titles = []
all_video_ids = []
# Get video info in uploads PL
while True:
    playlist = youtube.playlistItems().list(
        part="contentDetails",
        playlistId=all_uploads,
        maxResults=50,
        pageToken=pagetoken

    )

    pl_response = playlist.execute()

    #############
    # VIDEO IDS #
    #############
    batch_video_ids = []
    for item in pl_response['items']:
        # appennd to all_video_ids list
        all_video_ids.append(item['contentDetails']['videoId'])
        # appennd to batch_video_ids list which we will use to get video titles online 44
        batch_video_ids.append(item['contentDetails']['videoId'])

    ################
    # VIDEO TITLES #
    ################
    vid_title = youtube.videos().list(
        part='snippet',
        id=batch_video_ids
    )
    titleout = vid_title.execute()
    for title in titleout['items']:
        # append to the titles list
        titles.append(title['snippet']['title'])

    ############################
    # CHECKING NEXT PAGE TOKEN #
    ############################
    pagetoken = pl_response.get('nextPageToken')
    if pagetoken is None:
        break


# Create Excel spreadsheet
file = Workbook()
ws = file.active
ws.title = "" # Enter Worksheet Title
# Write to rows
for index, each_title in enumerate(titles):
    ws.cell(row=index + 1, column=1).value = each_title
    ws.row_dimensions[index + 1].height = 20
    ws.column_dimensions['A'].width = 100
for num, vid_id in enumerate(all_video_ids):
    ws.cell(row=num + 1, column=2).value = "https://www.youtube.com/watch?v=" + vid_id
ws.insert_rows(1)
ws.cell(row=1, column=1).value = "ALL VIDEO TITLES"
ws['A1'].font = Font(bold=True)
ws.cell(row=1, column=2).value = " VIDEO LINKS"
ws['B1'].font = Font(bold=True)
ws.column_dimensions['B'].width = 50

# Save spreadsheet & Name Spreadsheet
file.save('.xlsx')

